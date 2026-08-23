#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Render a HW06 case specification into every artefact that needs it.

    python scripts/render-cases.py --api 1

Reads scripts/cases/apiN_*.py (the single source of truth for that API's test
cases) and writes:

  collections/<name>.postman_collection.json   executable cases, foldered by
                                               coverage dimension
  testcases/<slug>_cases.json                  machine-readable export
  testcases/23127184_HW06_TestCases.xlsx       the API's sheet, refreshed
  reports/coverage_<slug>.md                   a coverage tally

Because everything derives from one module, a correction made during the audit
phase propagates to the collection, the workbook and the coverage tally in one
step - no chance of the Excel and the collection drifting apart.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import sys
from collections import Counter, OrderedDict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(Path(__file__).resolve().parent))

import postman_harness as H  # noqa: E402

CASE_MODULES = {
    1: "cases/api1_fr01_register.py",
}


def load_cases(api: int):
    rel = CASE_MODULES.get(api)
    if rel is None:
        raise SystemExit("No case specification registered for API %d." % api)
    path = Path(__file__).resolve().parent / rel
    spec = importlib.util.spec_from_file_location("hw06_cases_api%d" % api, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ---------------------------------------------------------------------------
# Postman collection
# ---------------------------------------------------------------------------

def build_request(case):
    return H.request(
        case.get("method", "POST"),
        case.get("path", "/api/register"),
        body=case.get("body"),
        raw_body=case.get("raw_body"),
        auth_var=case.get("auth_var"),
        extra_headers=case.get("headers"),
        content_type=case.get("content_type", "application/json"),
        description=describe(case),
    )


def describe(case):
    lines = [
        "**%s** - %s" % (case["id"], case["title"]),
        "",
        "| Field | Value |",
        "|---|---|",
        "| Dimension | %s |" % case["dim"],
        "| Parameter | %s |" % case.get("param", "-"),
        "| Partition | %s |" % case.get("partition", ""),
        "| Requirement | %s |" % case.get("rule", ""),
        "| Expected | %s |" % case.get("expected", ""),
    ]
    if case.get("gap"):
        lines += ["", "> **Specification gap:** %s" % case["gap"]]
    return "\n".join(lines)


def build_collection(module, only_ids=None, name_suffix="", extra_description=""):
    meta = module.META
    cases = module.CASES
    if only_ids is not None:
        cases = [c for c in cases if c["id"] in only_ids]
    items = []

    for folder_name, dimension, folder_desc in meta["folders"]:
        in_dimension = [c for c in cases if c["dim"] == dimension]
        if only_ids is not None and not in_dimension:
            continue
        subfolders = meta.get("subfolders", {}).get(folder_name)

        if subfolders:
            children = []
            for sub_name, param in subfolders:
                picked = [c for c in in_dimension if c.get("param") == param]
                if only_ids is not None and not picked:
                    continue
                children.append(H.folder(
                    sub_name,
                    [H.item("%s | %s" % (c["id"], c["title"]),
                            build_request(c), c["tests"], c.get("pre"))
                     for c in picked],
                    "%d case(s) on `%s`." % (len(picked), param),
                ))
            covered = {p for _, p in subfolders}
            leftovers = [c for c in in_dimension if c.get("param") not in covered]
            children.extend(
                H.item("%s | %s" % (c["id"], c["title"]),
                       build_request(c), c["tests"], c.get("pre"))
                for c in leftovers
            )
            items.append(H.folder(folder_name, children, folder_desc))
        else:
            items.append(H.folder(
                folder_name,
                [H.item("%s | %s" % (c["id"], c["title"]),
                        build_request(c), c["tests"], c.get("pre"))
                 for c in in_dimension],
                folder_desc,
            ))

    return H.collection(meta["collection_name"] + name_suffix,
                        meta["description"] + extra_description, items)


# ---------------------------------------------------------------------------
# CI regression gate, at case granularity
# ---------------------------------------------------------------------------
#
# Folder-level gating stopped being useful once the suite was written properly
# against the specification: the SUT fails cases scattered across almost every
# folder, so only one folder is entirely green and gating on it proves nothing.
#
# The gate is therefore a list of case IDs - every case that passes today - and
# a separate collection is rendered containing exactly those. The rule it
# encodes is "whatever works today must keep working", which is what a
# regression gate is for.

COLLECTION_NAMES = {1: "API1_FR01_Register"}


def passing_case_ids(module, report_path):
    """Case ids with zero failed assertions in the given Newman JSON report."""
    report = json.loads(report_path.read_text(encoding="utf-8"))
    failed = set()
    for failure in report["run"].get("failures", []):
        source = failure.get("source") or {}
        name = source.get("name", "")
        if " | " in name:
            failed.add(name.split(" | ")[0])

    executed = set()
    for execution in report["run"].get("executions", []):
        name = (execution.get("item") or {}).get("name", "")
        if " | " in name:
            executed.add(name.split(" | ")[0])

    known = {c["id"] for c in module.CASES}
    return sorted((executed & known) - failed)


def refresh_gate(module, api):
    name = COLLECTION_NAMES[api]
    report_path = ROOT / "reports" / ("%s.json" % name)
    if not report_path.exists():
        raise SystemExit(
            "No run report at %s - run the full suite first:\n"
            "    .\\scripts\\Invoke-ApiTests.ps1 -Api %d" % (report_path, api)
        )

    green = passing_case_ids(module, report_path)
    total = len(module.CASES)

    suite_path = ROOT / "config" / "ci-suite.json"
    suite = json.loads(suite_path.read_text(encoding="utf-8"))
    suite.setdefault("gate_cases", {})[name] = green
    suite.setdefault("gate", {}).pop(name, None)
    suite_path.write_text(json.dumps(suite, indent=2, ensure_ascii=False) + "\n",
                          encoding="utf-8")

    gate = build_collection(
        module, only_ids=set(green), name_suffix=" [CI gate]",
        extra_description=(
            "\n\n---\n\nCI REGRESSION GATE - generated, do not edit.\n\n"
            "Contains the %d of %d cases that passed the run recorded in "
            "reports/%s.json. The remaining %d fail because the SUT violates "
            "the specification; they are tracked in docs/bugs/BUG_REPORT.md, "
            "not here. Refresh with:\n\n"
            "    python scripts/render-cases.py --api %d --refresh-gate\n"
            % (len(green), total, name, total - len(green), api)
        ),
    )
    gate_path = ROOT / "collections" / ("%s_gate.postman_collection.json" % name)
    gate_path.write_text(json.dumps(gate, indent=2, ensure_ascii=False),
                         encoding="utf-8")
    return gate_path, green, total


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def render_excel(module):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not available - skipped the workbook.")
        return None

    path = ROOT / "testcases" / "23127184_HW06_TestCases.xlsx"
    if not path.exists():
        print("  workbook missing - run scripts/build-testcase-workbook.py first.")
        return None

    meta, cases = module.META, module.CASES
    wb = load_workbook(path)
    sheet_name = meta["sheet"][:31]
    if sheet_name not in wb.sheetnames:
        print("  sheet %r not found - skipped." % sheet_name)
        return None

    ws = wb[sheet_name]
    # Clear previous rows, keep the title block and the header row (rows 1-3).
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top = Alignment(vertical="top", wrap_text=True)
    gap_fill = PatternFill("solid", fgColor="FFF2CC")

    for offset, c in enumerate(cases):
        row = 4 + offset
        payload = c.get("raw_body")
        if payload is None and c.get("body") is not None:
            payload = json.dumps(c["body"], ensure_ascii=False)
        values = [
            c["id"],
            c["dim"],
            c.get("rule", ""),
            c["title"],
            "fixture in pre-request script" if c.get("pre") else "none",
            "%s %s" % (c.get("method", "POST"), c.get("path", "/api/register")),
            (payload or "")[:900],
            c.get("expected", ""),
            "AI-generated",
            "",                       # Audit label - phase 2
            "",                       # Audit reasoning - phase 2
            "",                       # Correction applied - phase 2
            "",                       # Actual result - phase 4
            "Not run",
            "",                       # Bug ID
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = top
        if c.get("gap"):
            ws.cell(row=row, column=8).fill = gap_fill
            note = ws.cell(row=row, column=11)
            note.value = "SPEC GAP: %s" % c["gap"]
            note.font = Font(italic=True, color="9C6500")

    # Refresh the counts on the summary sheet for this API.
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        for row in range(5, 9):
            if str(summary.cell(row=row, column=1).value or "").startswith(meta["sheet"][:8]):
                summary.cell(row=row, column=4, value=len(cases))     # AI-generated
                summary.cell(row=row, column=5, value=0)              # student-added
                break

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Coverage tally
# ---------------------------------------------------------------------------

def render_coverage(module):
    meta, cases = module.META, module.CASES
    by_dim = Counter(c["dim"] for c in cases)
    by_param = OrderedDict()
    for c in cases:
        by_param.setdefault(c.get("param", "-"), []).append(c)
    by_rule = Counter()
    for c in cases:
        for token in str(c.get("rule", "")).replace("/", ",").split(","):
            token = token.strip()
            if token:
                by_rule[token] += 1
    gaps = [c for c in cases if c.get("gap")]

    lines = [
        "# Coverage - %s" % meta["collection_name"],
        "",
        "Generated by `scripts/render-cases.py` from "
        "`scripts/cases/%s.py`. Do not edit by hand." % meta["slug"].replace("-", "_"),
        "",
        "| Metric | Value |",
        "|---|---:|",
        "| Endpoint | `%s` |" % meta["endpoint"],
        "| Requirement | %s (Pool %s) |" % (meta["requirement"], meta["pool"]),
        "| **Total test cases** | **%d** |" % len(cases),
        "| Brief's minimum | 35 |",
        "| Ratio to minimum | %.1fx |" % (len(cases) / 35.0),
        "",
        "## By coverage dimension",
        "",
        "| Dimension | Cases |",
        "|---|---:|",
    ]
    for dimension in ["Domain", "State", "Security", "Schema"]:
        lines.append("| %s | %d |" % (dimension, by_dim.get(dimension, 0)))
    lines += ["| **Total** | **%d** |" % len(cases), ""]

    lines += ["## By parameter", "", "| Parameter | Cases |", "|---|---:|"]
    for param, group in by_param.items():
        lines.append("| `%s` | %d |" % (param, len(group)))

    lines += ["", "## By requirement", "", "| Requirement | Cases |", "|---|---:|"]
    for rule, count in sorted(by_rule.items()):
        lines.append("| %s | %d |" % (rule, count))

    lines += [
        "",
        "## Specification gaps flagged (%d)" % len(gaps),
        "",
        "Cases where the specification is genuinely silent. Rather than invent "
        "an expectation, each asserts only what can be justified and records "
        "the ambiguity - these are the rows to resolve first in the audit phase.",
        "",
        "| Case | Gap |",
        "|---|---|",
    ]
    for c in gaps:
        lines.append("| %s | %s |" % (c["id"], c["gap"]))

    lines += ["", "## Case index", "", "| ID | Dimension | Requirement | Title |",
              "|---|---|---|---|"]
    for c in cases:
        lines.append("| %s | %s | %s | %s |"
                     % (c["id"], c["dim"], c.get("rule", ""), c["title"]))

    path = ROOT / "reports" / ("coverage_%s.md" % meta["slug"])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path, by_dim, gaps


# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--api", type=int, required=True, choices=sorted(CASE_MODULES))
    parser.add_argument("--refresh-gate", action="store_true",
                        help="rebuild the CI regression gate from the last run "
                             "report instead of re-rendering the artefacts")
    args = parser.parse_args()

    module = load_cases(args.api)
    meta, cases = module.META, module.CASES

    if args.refresh_gate:
        gate_path, green, total = refresh_gate(module, args.api)
        print("CI gate refreshed for API %d" % args.api)
        print("  green cases : %d of %d (%.0f%%)"
              % (len(green), total, 100.0 * len(green) / total))
        print("  gate suite  : %s" % gate_path.relative_to(ROOT))
        print("  case list   : config/ci-suite.json -> gate_cases")
        return

    ids = [c["id"] for c in cases]
    duplicates = [i for i, n in Counter(ids).items() if n > 1]
    if duplicates:
        raise SystemExit("Duplicate case ids: %s" % ", ".join(sorted(duplicates)))

    collection = build_collection(module)
    rendered = sum(count_items(f) for f in collection["item"])
    if rendered != len(cases):
        raise SystemExit(
            "Rendered %d requests but the spec holds %d cases - a case fell "
            "outside every folder filter." % (rendered, len(cases))
        )

    coll_path = ROOT / "collections" / (
        {1: "API1_FR01_Register"}[args.api] + ".postman_collection.json")
    coll_path.write_text(json.dumps(collection, indent=2, ensure_ascii=False),
                         encoding="utf-8")

    export = ROOT / "testcases" / ("%s_cases.json" % meta["slug"])
    export.write_text(json.dumps(
        [{k: v for k, v in c.items() if k not in ("tests", "pre")} for c in cases],
        indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    cov_path, by_dim, gaps = render_coverage(module)
    xlsx = render_excel(module)

    print("API %d - %s" % (args.api, meta["endpoint"]))
    print("  cases          : %d  (%.1fx the brief's minimum of 35)"
          % (len(cases), len(cases) / 35.0))
    for dimension in ["Domain", "State", "Security", "Schema"]:
        print("    %-10s %3d" % (dimension, by_dim.get(dimension, 0)))
    print("  spec gaps      : %d flagged for the audit phase" % len(gaps))
    print("  collection     : %s" % coll_path.relative_to(ROOT))
    print("  case export    : %s" % export.relative_to(ROOT))
    print("  coverage       : %s" % cov_path.relative_to(ROOT))
    if xlsx:
        print("  workbook sheet : %s -> %s" % (meta["sheet"], xlsx.relative_to(ROOT)))


def count_items(node):
    if "item" in node:
        return sum(count_items(child) for child in node["item"])
    return 1


if __name__ == "__main__":
    main()
