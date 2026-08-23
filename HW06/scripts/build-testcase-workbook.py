#!/usr/bin/env python3
"""Build the HW06 Excel test-case workbook skeleton.

Creates testcases/23127184_HW06_TestCases.xlsx with one sheet per API plus a
summary sheet. Columns match what the audit (phase 2) and extend (phase 3)
phases need, so a case can be filled in once and carried through.

Refuses to overwrite an existing workbook - once you start filling it in, the
workbook is the source of truth.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "testcases" / "23127184_HW06_TestCases.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("TC ID", 12),
    ("Dimension", 16),
    ("Requirement / SEC", 18),
    ("Title", 42),
    ("Precondition", 30),
    ("Request (method + path)", 34),
    ("Input / body", 40),
    ("Expected result", 40),
    ("Source", 12),
    ("Audit label", 13),
    ("Audit reasoning", 40),
    ("Correction applied", 34),
    ("Actual result", 34),
    ("Status", 11),
    ("Bug ID", 10),
]

APIS = [
    ("API1 FR-01 Register", "POST /api/register", "Pool A"),
    ("API2 FR-06 Product Detail", "GET /api/products/:id", "Pool A"),
    ("API3 FR-11 Order History", "GET /api/orders/my-orders", "Pool B"),
    ("API4 FR-13 Admin Dashboard", "GET /api/admin/orders", "Pool C"),
]

DIMENSIONS = "Domain,State transition,Security,Schema"
SOURCES = "AI-generated,Student-added"
LABELS = "VALID,INVALID,INCOMPLETE"
STATUSES = "Pass,Fail,Blocked,Not run"


def style_header(ws, row=3):
    for idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def add_validation(ws, col_letter, formula, first=4, last=400):
    dv = DataValidation(type="list", formula1='"%s"' % formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (col_letter, first, col_letter, last))


def build_api_sheet(wb, name, endpoint, pool):
    ws = wb.create_sheet(name[:31])
    ws["A1"] = name
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Endpoint: %s   |   %s   |   Target: >= 35 test cases" % (endpoint, pool)
    ws["A2"].font = Font(italic=True, color="595959")

    style_header(ws)
    add_validation(ws, "B", DIMENSIONS)
    add_validation(ws, "I", SOURCES)
    add_validation(ws, "J", LABELS)
    add_validation(ws, "N", STATUSES)
    return ws


def build_summary(wb):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "HW06 - API Testing | Test Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Student ID: 23127184   |   SUT: EShop (ttbhanh/eshop-sut)"
    ws["A2"].font = Font(italic=True, color="595959")

    headers = ["API", "Endpoint", "Pool", "AI-generated", "Student-added",
               "Total", "Executed", "Passed", "Failed", "Bugs found"]
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(title) + 4)
    ws.column_dimensions["B"].width = 30
    ws.row_dimensions[4].height = 28

    for offset, (name, endpoint, pool) in enumerate(APIS):
        row = 5 + offset
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=endpoint)
        ws.cell(row=row, column=3, value=pool)
        # Total = generated + added
        ws.cell(row=row, column=6, value="=D%d+E%d" % (row, row))
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = BORDER

    total_row = 5 + len(APIS)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(4, 11):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value="=SUM(%s5:%s%d)" % (letter, letter, total_row - 1)).font = Font(bold=True)
        ws.cell(row=total_row, column=col).border = BORDER
    ws.cell(row=total_row, column=1).border = BORDER

    ws.cell(row=total_row + 2, column=1,
            value="Keep these numbers identical to the ones in README.md and the main report.")
    ws.cell(row=total_row + 2, column=1).font = Font(italic=True, color="C00000")
    return ws


def main():
    if OUT.exists():
        print("%s already exists - not overwriting." % OUT.name)
        return 0

    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb)
    for name, endpoint, pool in APIS:
        build_api_sheet(wb, name, endpoint, pool)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("Wrote %s (%d sheets)." % (OUT.name, len(wb.sheetnames)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
